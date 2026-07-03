#!/usr/bin/env python3
"""
Vinculum Protocol — Universal Price Fetcher v9
===============================================
Reference price lookup for vault mint ratio calculations.
Run 2x/day. Target: within a few percent. Not for trading.

Built against: approved_assets_final.json (1001 assets, clean)

Cascade per asset:
  Tier 1 — CoinGecko  (batch 200 IDs per call, ~5 calls total)
  Tier 2 — DexScreener (contract address lookup for batch misses)
  Tier 3 — DexScreener (symbol search for remaining)
  Tier 4 — Community-token override (TigerOG, LionOG, FrogOG, WKC):
           DexScreener -> GeckoTerminal, against a dedicated price-source
           contract that may differ from the asset's display contract.

v9 vs v8 — TigerOG / LionOG / FrogOG price-source fix:
  - TigerOG, LionOG, and FrogOG each exist as a legacy token on BSC and a
    unified Axelar ITS token on Base. The registry displays them as the
    Base ITS token (registry "contract" field = ITS address, unchanged).
    But the ITS contracts on Base have thin/inconsistent liquidity, while
    the original legacy BSC contracts retain real market activity — nobody
    is required to bridge their legacy tokens, so BSC liquidity persists.
  - COMMUNITY_TOKENS now points ds_addr/gt_addr at the legacy BSC contract
    for these three (ds_chain/gt_network = "bsc") instead of the Base ITS
    address. Display chain in the registry is untouched — still Base.
  - Steps 2 and 3 now skip any asset whose key is in COMMUNITY_TOKENS.
    Previously, Step 2 (DexScreener by contract) would run against these
    three using their Base ITS contract *before* reaching Step 4, and if
    the thin ITS pool returned any price at all it would lock that in and
    Step 4's correct BSC-sourced override would never run. That's the
    likely cause of the intermittent (not clean-miss) coverage gap.
  - Fixed leftover "v6" strings in the log header, xlsx summary title, and
    HTTP User-Agent that were never updated in earlier version bumps.
  - No hardcoded prices anywhere.

Usage:
  python vinculum_price_fetcher_v9.py

Both files must be in the same folder:
  approved_assets_final.json
  vinculum_price_fetcher_v9.py

Outputs:
  vinculum_prices.json
  vinculum_prices.xlsx
  price_fetch_log.txt
"""

import json, os, sys, time
from datetime import datetime, timezone

try:
    import requests
except ImportError:
    sys.exit("Run: pip install requests openpyxl")

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# ── CONFIG ───────────────────────────────────────────────────────────────────
INPUT_FILE  = "approved_assets_final.json"
OUTPUT_JSON = "vinculum_prices.json"
OUTPUT_XLSX = "vinculum_prices.xlsx"
LOG_FILE    = "price_fetch_log.txt"

TIMEOUT_S   = 14
CG_BATCH    = 200
CG_PAUSE    = 4.0   # seconds between CoinGecko batch calls
DS_PAUSE    = 0.6   # DexScreener
GT_PAUSE    = 1.5   # GeckoTerminal
MAX_RETRY   = 3

NATIVE_CHAINS = {
    "Bitcoin","Dogecoin","Litecoin","DigiByte","Zcash","XRPL","Stellar","Cosmos"
}

DS_CHAIN = {
    "Ethereum":"ethereum","Base":"base","Arbitrum":"arbitrum",
    "Optimism":"optimism","Polygon":"polygon","BSC":"bsc",
    "Avalanche":"avalanche","Solana":"solana",
}

GT_NET = {
    "Ethereum":"eth","Base":"base","Arbitrum":"arbitrum",
    "Optimism":"optimism","Polygon":"polygon","BSC":"bsc",
    "Avalanche":"avax","Solana":"solana",
}

# Community tokens: no CoinGecko ID — priced via DexScreener first,
# then GeckoTerminal as fallback. Both sources tried for every token.
#
# TigerOG / LionOG / FrogOG display in the registry as their unified Axelar
# ITS token on Base (registry "contract" field), but that display contract
# often has thin/inconsistent liquidity. Each also has an original legacy
# token on BSC that nobody is ever required to bridge away from, so it
# keeps real market activity. ds_addr/gt_addr below deliberately point at
# the BSC legacy contract, not the Base display contract — this is a
# price-source override, keyed by contract address, independent of what
# the registry shows the user. See v9 changelog above.
COMMUNITY_TOKENS = {
    "TigerOG@Base": {
        # Price source: BNBTiger legacy contract on BSC (~$921K liquidity
        # on its main WBNB pool as of last check). Display stays Base ITS.
        "ds_chain":  "bsc",
        "ds_addr":   "0xAC68931B666E086E9de380CFDb0Fb5704a35dc2D",
        "gt_network":"bsc",
        "gt_addr":   "0xAC68931B666E086E9de380CFDb0Fb5704a35dc2D",
    },
    "LionOG@Base": {
        # Price source: BNBLion legacy contract on BSC (~$110K liquidity
        # as of last check). Display stays Base ITS.
        "ds_chain":  "bsc",
        "ds_addr":   "0xdA1689C5557564d06E2A546F8FD47350b9D44a73",
        "gt_network":"bsc",
        "gt_addr":   "0xdA1689C5557564d06E2A546F8FD47350b9D44a73",
    },
    "FrogOG@Base": {
        # Price source: BNBFrog legacy contract on BSC (~$108K liquidity
        # as of last check). Display stays Base ITS.
        "ds_chain":  "bsc",
        "ds_addr":   "0x64da67A12a46f1DDF337393e2dA12eD0A507Ad3D",
        "gt_network":"bsc",
        "gt_addr":   "0x64da67A12a46f1DDF337393e2dA12eD0A507Ad3D",
    },
    "WKC@Ethereum": {
        # Unchanged — not part of the Base ITS pattern.
        "ds_chain":  "ethereum",
        "ds_addr":   "0x6ec90334d89dbdc89e08a133271be3d104128edb",
        "gt_network":"eth",
        "gt_addr":   "0x6ec90334d89dbdc89e08a133271be3d104128edb",
    },
}

fh = None

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    if fh and not fh.closed:
        fh.write(line + "\n"); fh.flush()

def now_ts():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def fetch(url, params=None):
    for attempt in range(1, MAX_RETRY + 1):
        try:
            r = requests.get(url, params=params, timeout=TIMEOUT_S,
                             headers={"Accept":"application/json",
                                      "User-Agent":"vinculum-price-fetcher/9"})
            if r.status_code == 200:
                return r.json()
            if r.status_code == 429:
                wait = 12 * attempt
                log(f"  429 rate-limit -- waiting {wait}s")
                time.sleep(wait); continue
            if r.status_code == 404:
                return None
        except Exception as e:
            log(f"  fetch error (attempt {attempt}): {e}")
        time.sleep(3 * attempt)
    return None

# ── TIER 1: COINGECKO BATCH ──────────────────────────────────────────────────
def cg_batch(id_list):
    """Fetch prices for a list of CoinGecko IDs. Returns {cg_id: usd}."""
    result = {}
    for i in range(0, len(id_list), CG_BATCH):
        batch = id_list[i:i+CG_BATCH]
        data = fetch("https://api.coingecko.com/api/v3/simple/price",
                     params={"ids": ",".join(batch), "vs_currencies": "usd"})
        if data:
            for cg_id, vals in data.items():
                if vals.get("usd"):
                    result[cg_id] = float(vals["usd"])
        if i + CG_BATCH < len(id_list):
            time.sleep(CG_PAUSE)
    return result

# ── TIER 2: DEXSCREENER BY CONTRACT ──────────────────────────────────────────
def ds_by_contract(chain, contract):
    ds_chain = DS_CHAIN.get(chain)
    if not ds_chain or not contract:
        return None
    # Skip placeholder values
    if any(contract.upper().startswith(x) for x in
           ("MISSING","PENDING","REVIEW","NATIVE","DEADBEEF")):
        return None
    data = fetch(f"https://api.dexscreener.com/latest/dex/tokens/{contract}")
    if not data or not data.get("pairs"):
        return None
    pairs = [p for p in data["pairs"]
             if p.get("chainId") == ds_chain
             and float(p.get("priceUsd") or 0) > 0]
    if not pairs:
        return None
    best = max(pairs, key=lambda p: float((p.get("liquidity") or {}).get("usd") or 0))
    return float(best["priceUsd"])

# ── TIER 3: DEXSCREENER SYMBOL SEARCH ────────────────────────────────────────
def ds_by_search(symbol, chain):
    ds_chain = DS_CHAIN.get(chain)
    if not ds_chain:
        return None
    data = fetch("https://api.dexscreener.com/latest/dex/search",
                 params={"q": symbol})
    if not data or not data.get("pairs"):
        return None
    pairs = [p for p in data["pairs"]
             if p.get("chainId") == ds_chain
             and (p.get("baseToken") or {}).get("symbol","").upper() == symbol.upper()
             and float(p.get("priceUsd") or 0) > 0
             and float((p.get("liquidity") or {}).get("usd") or 0) > 200]
    if not pairs:
        return None
    best = max(pairs, key=lambda p: float((p.get("liquidity") or {}).get("usd") or 0))
    return float(best["priceUsd"])

# ── TIER 4: GECKO TERMINAL ────────────────────────────────────────────────────
def gt_by_contract(network, address):
    data = fetch(f"https://api.geckoterminal.com/api/v2/networks/{network}/tokens/{address}")
    if not data:
        return None
    p = (data.get("data") or {}).get("attributes", {}).get("price_usd")
    try:
        v = float(p) if p else None
        return v if v and v > 0 else None
    except (ValueError, TypeError):
        return None

# ── FORMAT ───────────────────────────────────────────────────────────────────
def fmt(usd):
    if usd is None: return "--"
    if usd >= 1000:       return f"${usd:,.2f}"
    if usd >= 1:          return f"${usd:.4f}"
    if usd >= 0.0001:     return f"${usd:.6f}"
    if usd >= 0.000000001: return f"${usd:.12f}"
    return f"${usd:.2e}"

# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    global fh

    if not os.path.exists(INPUT_FILE):
        sys.exit(f"'{INPUT_FILE}' not found in current folder.")

    assets = json.load(open(INPUT_FILE, encoding="utf-8"))
    fh = open(LOG_FILE, "w", encoding="utf-8")
    ts  = now_ts()

    log("=== Vinculum Price Fetcher v9 ===")
    log(f"Assets: {len(assets)} | Purpose: vault reference rates (2x/day)")
    log(f"Input:  {INPUT_FILE}")
    log("")

    prices = {}

    # ── STEP 1: COINGECKO BATCH ──────────────────────────────────────────────
    log("Step 1: CoinGecko batch...")
    cg_map = {}
    for a in assets:
        cgid = a.get("coingecko_id","")
        if cgid:
            cg_map.setdefault(cgid, []).append((a["symbol"], a["chain"]))

    log(f"  {len(cg_map)} unique CoinGecko IDs across {len(assets)} assets")
    cg_prices = cg_batch(list(cg_map.keys()))
    applied = 0
    for cgid, asset_list in cg_map.items():
        usd = cg_prices.get(cgid)
        if usd:
            for sym, chain in asset_list:
                key = f"{sym}@{chain}"
                if key not in prices:
                    prices[key] = {"usd": usd, "source": "CoinGecko", "ts": ts}
                    applied += 1
    missed = len(cg_map) - len(cg_prices)
    log(f"  Got {len(cg_prices)} prices -> {applied} assets covered ({missed} IDs returned no price)")

    # ── STEP 2: DEXSCREENER BY CONTRACT (for batch misses) ───────────────────
    log("Step 2: DexScreener (contract address)...")
    ds_needed = [a for a in assets
                 if f"{a['symbol']}@{a['chain']}" not in prices
                 and f"{a['symbol']}@{a['chain']}" not in COMMUNITY_TOKENS
                 and a["chain"] in DS_CHAIN
                 and a.get("contract","")]
    log(f"  {len(ds_needed)} assets to try...")
    ds_applied = 0
    for a in ds_needed:
        key = f"{a['symbol']}@{a['chain']}"
        usd = ds_by_contract(a["chain"], a.get("contract",""))
        if usd:
            prices[key] = {"usd": usd, "source": "DexScreener", "ts": ts}
            ds_applied += 1
        time.sleep(DS_PAUSE)
    log(f"  DexScreener contract: {ds_applied} prices found")

    # ── STEP 3: DEXSCREENER SYMBOL SEARCH (still missing) ────────────────────
    log("Step 3: DexScreener symbol search...")
    ds_search_needed = [a for a in assets
                        if f"{a['symbol']}@{a['chain']}" not in prices
                        and f"{a['symbol']}@{a['chain']}" not in COMMUNITY_TOKENS
                        and a["chain"] in DS_CHAIN]
    log(f"  {len(ds_search_needed)} assets to try...")
    ds_search_applied = 0
    for a in ds_search_needed:
        key = f"{a['symbol']}@{a['chain']}"
        usd = ds_by_search(a["symbol"], a["chain"])
        if usd:
            prices[key] = {"usd": usd, "source": "DexScreener_search", "ts": ts}
            ds_search_applied += 1
        time.sleep(DS_PAUSE)
    log(f"  DexScreener search: {ds_search_applied} prices found")

    # ── STEP 4: COMMUNITY TOKENS (DexScreener first, GeckoTerminal fallback) ──
    log("Step 4: Community tokens (DexScreener -> GeckoTerminal)...")
    comm_applied = 0
    for key, cfg in COMMUNITY_TOKENS.items():
        if key in prices:
            log(f"  {key}: already priced")
            continue

        usd = None
        source = None

        # Try DexScreener first
        ds_data = fetch(f"https://api.dexscreener.com/latest/dex/tokens/{cfg['ds_addr']}")
        if ds_data and ds_data.get("pairs"):
            pairs = [p for p in ds_data["pairs"]
                     if p.get("chainId") == cfg["ds_chain"]
                     and float(p.get("priceUsd") or 0) > 0]
            if pairs:
                best = max(pairs, key=lambda p: float((p.get("liquidity") or {}).get("usd") or 0))
                usd = float(best["priceUsd"])
                source = "DexScreener"
        time.sleep(DS_PAUSE)

        # GeckoTerminal fallback if DexScreener found nothing
        if not usd:
            usd = gt_by_contract(cfg["gt_network"], cfg["gt_addr"])
            if usd:
                source = "GeckoTerminal"
            time.sleep(GT_PAUSE)

        if usd:
            prices[key] = {"usd": usd, "source": source, "ts": ts}
            comm_applied += 1
            log(f"  {key}: ${usd} [{source}]")
        else:
            prices[key] = {"usd": None, "source": "none", "ts": ts,
                           "error": "no price on DexScreener or GeckoTerminal — low/no liquidity"}
            log(f"  {key}: no price found on either source")

    log(f"  Community tokens: {comm_applied} priced")

    # ── STEP 5: MARK REMAINING ────────────────────────────────────────────────
    no_price = []
    for a in assets:
        key = f"{a['symbol']}@{a['chain']}"
        if key not in prices:
            prices[key] = {"usd": None, "source": "none", "ts": ts,
                           "error": "no price source resolved"}
            no_price.append(key)

    # ── SUMMARY ──────────────────────────────────────────────────────────────
    total  = len(assets)
    priced = sum(1 for a in assets
                 if prices.get(f"{a['symbol']}@{a['chain']}",{}).get("usd") is not None)
    from collections import Counter
    sources = Counter(prices[f"{a['symbol']}@{a['chain']}"]["source"]
                      for a in assets
                      if prices.get(f"{a['symbol']}@{a['chain']}",{}).get("usd"))

    log("")
    log("=== FINAL SUMMARY ===")
    log(f"Total assets:  {total}")
    log(f"Priced:        {priced} ({100*priced/total:.1f}%)")
    log(f"No price:      {total - priced}")
    log("")
    for src, cnt in sources.most_common():
        log(f"  {src:30} {cnt}")
    if no_price:
        log(f"\nNot priced ({len(no_price)}):")
        for k in no_price:
            log(f"  {k}")

    # ── SAVE JSON ─────────────────────────────────────────────────────────────
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(prices, f, indent=1)
    log(f"\nJSON saved: {OUTPUT_JSON}")

    # ── SAVE XLSX ─────────────────────────────────────────────────────────────
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Prices"
        headers = ["Symbol","Chain","Price (USD)","Source","Timestamp",
                   "CoinGecko ID","Contract","Note"]
        ws.append(headers)
        hfill = PatternFill("solid", start_color="1A1A2E")
        hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        for c in ws[1]:
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        src_fills = {
            "CoinGecko":          PatternFill("solid", start_color="E3F2FD"),
            "DexScreener":        PatternFill("solid", start_color="EDE7F6"),
            "DexScreener_search": PatternFill("solid", start_color="F3E5F5"),
            "GeckoTerminal":      PatternFill("solid", start_color="FFF3E0"),
            "none":               PatternFill("solid", start_color="FFEBEE"),
        }

        for a in assets:
            key = f"{a['symbol']}@{a['chain']}"
            val = prices.get(key, {})
            usd = val.get("usd")
            ws.append([
                a["symbol"], a["chain"], fmt(usd),
                val.get("source",""), val.get("ts",""),
                a.get("coingecko_id",""), a.get("contract",""),
                val.get("error","")
            ])
            r = ws.max_row
            sfill = src_fills.get(val.get("source","none"),
                                   PatternFill("solid", start_color="FFFFFF"))
            for col in range(1, 9):
                ws.cell(row=r, column=col).font = Font(name="Arial", size=9)
            ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=1).fill = sfill
            ws.cell(row=r, column=4).fill = sfill

        for i, w in enumerate([14, 11, 16, 20, 20, 28, 46, 36], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        # Summary sheet
        ss = wb.create_sheet("Summary")
        ss.append(["Vinculum Price Fetcher v9 -- Run Summary"])
        ss.append(["Run time", ts])
        ss.append(["Input file", INPUT_FILE])
        ss.append(["Total assets", total])
        ss.append(["Priced", priced])
        ss.append(["Coverage", f"{100*priced/total:.1f}%"])
        ss.append([])
        ss.append(["Source", "Count"])
        for src, cnt in sources.most_common():
            ss.append([src, cnt])
        if no_price:
            ss.append([])
            ss.append(["Not priced", ""])
            for k in no_price:
                ss.append(["", k])
        ss["A1"].font = Font(name="Arial", bold=True, size=12)
        ss.column_dimensions["A"].width = 30
        ss.column_dimensions["B"].width = 50

        wb.save(OUTPUT_XLSX)
        log(f"Excel saved: {OUTPUT_XLSX}")

    except ImportError:
        log("openpyxl not installed -- xlsx skipped. JSON output is complete.")

    log(f"\nDone. {priced}/{total} assets priced ({100*priced/total:.1f}% coverage).")
    fh.close()


if __name__ == "__main__":
    main()
